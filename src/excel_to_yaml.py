"""
Excel-to-YAML converter for automated test cases.
Reads test cases from an Excel file and generates individual YAML test files.

Expected Excel format:
  Column A: Test Name  (or Test ID like TC_AUTH_01)
  Column B: Description
  Column C: Steps (numbered, one per line)
  Column D: Expected Results (numbered, one per line)
  Column E: Tags (comma-separated, optional)

Each row = one test case → one YAML file in the tests/ directory.

Usage:
  python excel_to_yaml.py <excel_file> [output_dir]   # convert all
  python excel_to_yaml.py <excel_file> --list          # list tests
  python excel_to_yaml.py <excel_file> --select 1,3,5  # convert specific rows
  python excel_to_yaml.py --sample                     # create a sample template
"""

import os
import re
import sys
import yaml
from openpyxl import Workbook, load_workbook


def parse_steps(steps_text):
    """Parse a numbered steps column into a structured step list.

    Handles format like:
      1. Open admin console:
         admin-tool login
      2. List resources:
         admin-tool list --all
    """
    steps = []
    current_step = None
    current_commands = []

    for line in steps_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue

        step_match = re.match(r'^(\d+)\.\s*(.*)', line)
        if step_match:
            if current_step:
                steps.append(_build_step(current_step, current_commands))
            current_step = step_match.group(2).rstrip(':')
            current_commands = []
        else:
            current_commands.append(line)

    if current_step:
        steps.append(_build_step(current_step, current_commands))

    # Resolve "from Step X" references — copy commands from referenced step
    for i, step in enumerate(steps):
        if not step["responses"]:
            ref_match = re.search(r'(?:from|of)\s+step\s+(\d+)', step["name"], re.IGNORECASE)
            if ref_match:
                ref_idx = int(ref_match.group(1)) - 1
                if 0 <= ref_idx < len(steps) and steps[ref_idx]["responses"]:
                    step["responses"] = list(steps[ref_idx]["responses"])

    return steps


def _build_step(step_name, commands):
    """Build a step dict from name and commands."""
    responses = []
    for cmd in commands:
        cmd = cmd.strip()
        if cmd and not cmd.startswith('#') and not cmd.startswith('NOTE:'):
            responses.append(cmd)

    wait = 5
    name_lower = step_name.lower()
    if any(k in name_lower for k in ['wait', 'restart', 'kill', 'long']):
        wait = 15
    if any(k in name_lower for k in ['concurrent', 'load', 'parallel']):
        wait = 30

    return {
        "name": step_name,
        "command": "bash",
        "responses": responses,
        "wait": wait,
        "criteria": {
            "error_keywords": ["ERROR", "FAILED", "denied"]
        }
    }


def parse_expected_results(expected_text):
    """Parse the expected results column into a list of strings."""
    results = []
    for line in expected_text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        line = re.sub(r'^\d+\.\s*', '', line)
        results.append(line)
    return results


def apply_expected_to_steps(steps, expected_results):
    """Apply expected results as pass/fail criteria hints to corresponding steps."""
    for result in expected_results:
        result_lower = result.lower()

        if 'fail with error' in result_lower or 'commands fail' in result_lower:
            match = re.match(r'^(\d+)', result)
            if match:
                idx = int(match.group(1)) - 1
                if idx < len(steps):
                    steps[idx]["criteria"]["expect_failure"] = True
                    steps[idx]["criteria"]["expected_output"] = "ERROR"

        for keyword in ['OK', 'SUCCESS', 'COMPLETE']:
            if keyword in result:
                for step in reversed(steps):
                    if 'expected_output' not in step.get("criteria", {}):
                        step["criteria"]["expected_output"] = keyword
                        break

    return steps


def excel_to_yaml(excel_path, output_dir="tests", selected_indices=None):
    """Convert Excel test cases to individual YAML files.

    Args:
        excel_path:       Path to the Excel file.
        output_dir:       Directory to write YAML files into.
        selected_indices: List of 1-based row numbers to convert (None = all).
    """
    wb = load_workbook(excel_path, read_only=True)
    ws = wb.active

    os.makedirs(output_dir, exist_ok=True)
    generated_files = []

    for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=1):
        if not row or not row[0]:
            continue

        if selected_indices and idx not in selected_indices:
            continue

        col_a = str(row[0]).strip()
        # If col A looks like a Test ID (e.g. TC_AUTH_01), use ID+Name+Description layout
        if re.match(r'^[A-Z]{2,}_[A-Z]+_\d+$', col_a):
            test_id    = col_a
            test_name  = str(row[1]).strip() if len(row) > 1 and row[1] else test_id
            description= str(row[2]).strip() if len(row) > 2 and row[2] else ""
            steps_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            expected_text = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            tags_text  = str(row[5]).strip() if len(row) > 5 and row[5] else ""
        else:
            test_id    = None
            test_name  = col_a
            description= str(row[1]).strip() if len(row) > 1 and row[1] else ""
            steps_text = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            expected_text = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            tags_text  = str(row[4]).strip() if len(row) > 4 and row[4] else ""

        if not steps_text:
            print(f"  Skipping row {idx}: no steps defined")
            continue

        steps = parse_steps(steps_text)
        expected_results = parse_expected_results(expected_text) if expected_text else []

        if expected_results:
            steps = apply_expected_to_steps(steps, expected_results)

        tags = [t.strip() for t in tags_text.split(',') if t.strip()] if tags_text else []

        test_case = {
            "name": test_name,
            "description": description,
            "tags": tags,
            "steps": steps,
        }
        if expected_results:
            test_case["expected_result"] = "; ".join(expected_results)

        if test_id:
            filename = f"{test_id}.yaml"
        else:
            safe_name = re.sub(r'[^a-zA-Z0-9_]', '_', test_name).strip('_')
            filename = f"{safe_name}.yaml"
        filepath = os.path.join(output_dir, filename)

        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(f"# {filename.replace('.yaml', '')}: {test_name}\n")
            f.write(f"# {description}\n\n")
            yaml.dump(test_case, f, default_flow_style=False, sort_keys=False, allow_unicode=True)

        generated_files.append(filepath)
        print(f"  Generated: {filepath}")

    wb.close()
    return generated_files


def create_sample_excel(output_path="config/test_cases_sample.xlsx"):
    """Create a sample Excel template showing the expected format."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    headers = ["Test Name", "Description", "Steps", "Expected Results", "Tags"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # --- Sample test 1 ---
    ws.cell(row=2, column=1, value="Validate service behavior when dependency is stopped")
    ws.cell(row=2, column=2, value="Validates that the service correctly handles a missing dependency")
    ws.cell(row=2, column=3, value="""1. Stop the dependency service:
   service-ctl stop dependency-svc
2. Attempt operations (expect failure):
   admin-tool run-operation --type basic
3. Restart the dependency service:
   service-ctl start dependency-svc
4. Verify operations succeed after restart:
   admin-tool run-operation --type basic""")
    ws.cell(row=2, column=4, value="""2. Commands fail with ERROR
4. SUCCESS""")
    ws.cell(row=2, column=5, value="resilience, dependency, service")

    # --- Sample test 2 ---
    ws.cell(row=3, column=1, value="Validate resource generation and lifecycle")
    ws.cell(row=3, column=2, value="Validates create → list → deactivate → delete resource flow")
    ws.cell(row=3, column=3, value="""1. Generate resource and extract ID:
   admin-tool generate-resource
2. List resources and verify ID present:
   admin-tool list-resources
3. Deactivate resource:
   admin-tool deactivate {{resource_id}}
4. Delete resource:
   admin-tool delete {{resource_id}}""")
    ws.cell(row=3, column=4, value="""1. SUCCESS with resource_id
2. resource_id in listing
3. SUCCESS
4. SUCCESS""")
    ws.cell(row=3, column=5, value="lifecycle, resource, variable-chaining")

    os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
    wb.save(output_path)
    print(f"Sample Excel created: {output_path}")
    return output_path


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage:")
        print("  python excel_to_yaml.py <excel_file> [output_dir]   # Convert all tests")
        print("  python excel_to_yaml.py <excel_file> --list          # List tests in Excel")
        print("  python excel_to_yaml.py <excel_file> --select 1,3,5  # Convert specific tests")
        print("  python excel_to_yaml.py --sample                     # Create sample template")
        sys.exit(1)

    if sys.argv[1] == "--sample":
        create_sample_excel()
        sys.exit(0)

    excel_path = sys.argv[1]
    if not os.path.exists(excel_path):
        print(f"ERROR: File not found: {excel_path}")
        sys.exit(1)

    if len(sys.argv) > 2 and sys.argv[2] == "--list":
        wb = load_workbook(excel_path, read_only=True)
        ws = wb.active
        print(f"\nTests in: {excel_path}\n")
        print(f"{'#':<4} {'Name':<55} {'Tags'}")
        print("-" * 80)
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 1):
            name = row[0] if row[0] else "(empty)"
            tags = row[4] if len(row) > 4 and row[4] else ""
            print(f"{idx:<4} {str(name):<55} {tags}")
        wb.close()
        print(f"\nUse --select 1,3,5 to convert specific tests.")
        sys.exit(0)

    output_dir = "tests"
    selected = None

    remaining = sys.argv[2:]
    if "--select" in remaining:
        idx = remaining.index("--select")
        selected = [int(x.strip()) for x in remaining[idx + 1].split(",")]
        remaining = remaining[:idx] + remaining[idx + 2:]
    if remaining:
        output_dir = remaining[0]

    files = excel_to_yaml(excel_path, output_dir, selected)
    print(f"\nGenerated {len(files)} YAML file(s) in {output_dir}/")
